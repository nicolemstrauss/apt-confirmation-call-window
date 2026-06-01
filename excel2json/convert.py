#!/usr/bin/env python3
"""Convert a populated practice-time-window Excel workbook into per-practice JSON files."""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

TIME_RE = re.compile(r"^([01]\d|2[0-3]):[0-5]\d$")

DAY_COLUMNS = {
    "Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
    "Friday": 4, "Saturday": 5, "Sunday": 6,
}


def validate_time(value: str, phone: str, day: str, errors: list):
    """Validate a single HH:MM value is in 24-hour format."""
    if not TIME_RE.match(value):
        errors.append(f"  {phone} | {day}: '{value}' is not valid military time (expected HH:MM, 00:00-23:59)")


def parse_windows(cell_value: str, dow: int, phone: str, day: str, errors: list) -> list[dict]:
    """Parse 'HH:MM-HH:MM, HH:MM-HH:MM' into window dicts."""
    if not cell_value or not str(cell_value).strip():
        return []
    windows = []
    for rng in str(cell_value).split(","):
        rng = rng.strip()
        if not rng:
            continue
        parts = rng.split("-")
        if len(parts) != 2:
            errors.append(f"  {phone} | {day}: '{rng}' is not a valid range (expected HH:MM-HH:MM)")
            continue
        start, end = parts[0].strip(), parts[1].strip()
        validate_time(start, phone, day, errors)
        validate_time(end, phone, day, errors)
        windows.append({"dow": dow, "start": start, "end": end})
    return windows


def read_excel(path: str) -> dict:
    """Return {phone: json_body} from the workbook."""
    wb = openpyxl.load_workbook(path)
    ws = wb["config"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    results = {}
    errors = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        record = dict(zip(headers, row))
        phone = str(record.pop("phone")).strip()

        body = {
            "tz": str(record["tz"]),
            "max_days": int(record["max_days"]),
            "attempts_per_day": int(record["attempts_per_day"]),
        }

        windows = []
        for day_name, dow in DAY_COLUMNS.items():
            windows.extend(parse_windows(record.get(day_name), dow, phone, day_name, errors))
        body["windows"] = windows

        body["alternate_ivr_behavior"] = bool(record.get("alternate_ivr_behavior"))
        if record.get("alternate_ivr_description"):
            body["alternate_ivr_description"] = str(record["alternate_ivr_description"])

        if phone in results:
            errors.append(f"  Duplicate phone number: {phone}")
        results[phone] = body

    if errors:
        print("ERROR: Invalid time values found:\n")
        print("\n".join(errors))
        print("\nAll times must be in 24-hour military format (HH:MM), e.g. 08:00, 13:00, 17:30")
        sys.exit(1)

    return results


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", help="Path to the populated Excel workbook")
    parser.add_argument("-o", "--output-dir", default="output", help="Directory for JSON files (default: output)")
    args = parser.parse_args()

    configs = read_excel(args.excel)
    if not configs:
        print("No practice rows found.")
        raise SystemExit(1)

    out = Path(args.output_dir)
    out.mkdir(parents=True, exist_ok=True)

    for phone, body in configs.items():
        filename = phone.lstrip("+") + ".json"
        (out / filename).write_text(json.dumps(body, indent=2) + "\n")
        print(f"Wrote {out / filename}")

    print(f"\n{len(configs)} file(s) generated in {out}/")


if __name__ == "__main__":
    main()
