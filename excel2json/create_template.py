#!/usr/bin/env python3
"""Generate a blank practice-time-window Excel template."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "phone", "tz", "max_days", "attempts_per_day",
    "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
    "alternate_ivr_behavior", "alternate_ivr_description",
]

EXAMPLE_ROW = [
    "+12143309221", "America/Chicago", 2, 1,
    "08:00-12:00, 13:00-17:00",  # Mon
    "08:00-12:00, 13:00-17:00",  # Tue
    "08:00-12:00, 13:00-17:00",  # Wed
    "08:00-12:00, 13:00-17:00",  # Thu
    "08:00-12:00, 13:00-16:00",  # Fri
    "", "",  # Sat, Sun
    True,
    "When encountering IVR, do not engage with the IVR system and instead wait for a human to answer.",
]

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "config"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col, val in enumerate(EXAMPLE_ROW, 1):
        ws.cell(row=2, column=col, value=val)

    # timezone dropdown
    tz_list = '"America/New_York,America/Chicago,America/Denver,America/Los_Angeles,America/Phoenix"'
    tz_dv = DataValidation(type="list", formula1=tz_list, allow_blank=False)
    tz_dv.error = "Please select a valid timezone"
    tz_dv.errorTitle = "Invalid timezone"
    tz_dv.prompt = "Select a timezone"
    tz_dv.promptTitle = "Timezone"
    tz_dv.showDropDown = False
    tz_col = "B"
    tz_dv.add(f"{tz_col}2:{tz_col}1048576")
    ws.add_data_validation(tz_dv)

    # input prompts on day-of-week columns
    day_dv = DataValidation()
    day_dv.prompt = "Use 24-hour time: HH:MM-HH:MM\nMultiple windows: 08:00-12:00, 13:00-17:00\nLeave blank for no windows."
    day_dv.promptTitle = "Time Windows"
    day_dv.showInputMessage = True
    day_cols = "EFGHIJK"  # Monday through Sunday
    for letter in day_cols:
        day_dv.add(f"{letter}2:{letter}1048576")
    ws.add_data_validation(day_dv)

    # auto-size columns
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Unlock data cells (rows 2+, columns A-M) so customers can enter data
    from openpyxl.styles import Protection
    unlocked = Protection(locked=False)
    for row in range(2, 1001):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row, column=col).protection = unlocked

    # Protect sheet — prevents adding/deleting columns/rows beyond data area
    ws.protection.sheet = True
    ws.protection.insert_columns = True
    ws.protection.delete_columns = True
    ws.protection.insert_rows = False
    ws.protection.delete_rows = False
    ws.protection.format_cells = False

    path = "practice_time_window_template.xlsx"
    wb.save(path)
    print(f"Template saved to {path}")

if __name__ == "__main__":
    main()
